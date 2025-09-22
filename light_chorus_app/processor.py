"""Core MIDI parsing and spreadsheet generation logic for the Light Chorus app."""
from __future__ import annotations

from dataclasses import dataclass
from fractions import Fraction
from typing import Dict, List, Optional, Sequence, Tuple

import mido
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# MIDI note number associated with primer tone Short0 (see project documentation)
# Default so that Short0.mp3 corresponds to MIDI note 36 (C2 in scientific pitch notation)
SHORT_SAMPLE_BASE_NOTE = 36
# Aligns computed measure numbers with the shared event recipe spreadsheet
MEASURE_NUMBER_OFFSET = -2

# Row ordering and styling metadata for the nine Light Chorus parts
PART_DEFINITIONS: Sequence[Tuple[str, str]] = (
    ("Soprano 3", "#00a651"),  # Green
    ("Soprano 4", "#cc00ff"),  # Magenta
    ("Soprano 5", "#ff8300"),  # Orange
    ("Alto 3", "#1f6cff"),     # Blue
    ("Alto 4", "#d12c32"),     # Red
    ("Alto 5", "#00c8ff"),     # Cyan
    ("Tenor 2", "#ffc000"),    # Yellow
    ("Baritone 2", "#ff6fae"), # Pink
    ("Bass 3", "#7d4bff"),     # Purple
)

PREFERRED_ACCIDENTAL_MAP = {
    "C#": "Db",
    "D#": "Eb",
    "F#": "Gb",
    "G#": "Ab",
    "A#": "Bb",
}


@dataclass(frozen=True)
class NoteEntry:
    """Container for a MIDI pitch translated to project terminology."""

    midi: int
    note_name: str
    short_sample: str

    @property
    def sample_path(self) -> str:
        return f"primerTones/short{self.short_sample}.mp3"


@dataclass
class EventColumn:
    """All Light Chorus information for a single onset instant."""

    number: int
    tick: int
    measure: int
    position_label: str
    part_notes: Dict[str, List[NoteEntry]]


@dataclass
class SignatureSegment:
    """Represents a span of ticks that share the same time signature."""

    start_tick: int
    numerator: int
    denominator: int
    beat_ticks: int
    measure_ticks: int
    measures_before: int
    end_tick: Optional[int] = None

    def contains(self, tick: int) -> bool:
        upper_ok = self.end_tick is None or tick < self.end_tick
        return self.start_tick <= tick and upper_ok


class TimeSignatureMap:
    """Efficiently converts absolute ticks to measure and beat positions."""

    def __init__(self, midi_file: mido.MidiFile) -> None:
        self.ticks_per_beat = midi_file.ticks_per_beat
        self.segments: List[SignatureSegment] = self._build_segments(midi_file.tracks[0])
        if not self.segments:
            raise ValueError("No time signature metadata found in track 0")

    def _build_segments(self, track: mido.MidiTrack) -> List[SignatureSegment]:
        segments: List[SignatureSegment] = []
        abs_tick = 0
        raw_signatures: List[Tuple[int, int, int]] = []
        for message in track:
            abs_tick += message.time
            if message.type == "time_signature":
                raw_signatures.append((abs_tick, message.numerator, message.denominator))
        if not raw_signatures or raw_signatures[0][0] != 0:
            # Assume an initial 4/4 if not explicitly stated at tick 0
            raw_signatures.insert(0, (0, 4, 4))

        raw_signatures.sort(key=lambda item: item[0])
        for index, (start_tick, numerator, denominator) in enumerate(raw_signatures):
            next_start = raw_signatures[index + 1][0] if index + 1 < len(raw_signatures) else None
            beat_ticks_fraction = Fraction(self.ticks_per_beat) * Fraction(4, denominator)
            beat_ticks = int(beat_ticks_fraction)
            measure_ticks = beat_ticks * numerator
            segment = SignatureSegment(
                start_tick=start_tick,
                numerator=numerator,
                denominator=denominator,
                beat_ticks=beat_ticks,
                measure_ticks=measure_ticks,
                measures_before=0,
                end_tick=next_start,
            )
            segments.append(segment)

        cumulative_measures = 0
        for segment in segments:
            segment.measures_before = cumulative_measures
            if segment.end_tick is None:
                continue
            ticks_in_span = segment.end_tick - segment.start_tick
            measures_in_span = ticks_in_span / segment.measure_ticks
            cumulative_measures += int(round(measures_in_span))
        return segments

    def measure_position(self, tick: int) -> Tuple[int, str]:
        segment = next((seg for seg in reversed(self.segments) if seg.contains(tick)), None)
        if segment is None:
            segment = self.segments[0]
        ticks_into_segment = tick - segment.start_tick
        measure_offset = ticks_into_segment // segment.measure_ticks
        ticks_into_measure = ticks_into_segment % segment.measure_ticks
        beat_index = ticks_into_measure // segment.beat_ticks
        leftover_ticks = ticks_into_measure % segment.beat_ticks
        measure_number = (
            segment.measures_before + measure_offset + 1 + MEASURE_NUMBER_OFFSET
        )
        beat_label = self._format_beat_label(int(beat_index) + 1, segment.numerator, leftover_ticks, segment.beat_ticks)
        return measure_number, beat_label

    @staticmethod
    def _format_beat_label(beat_number: int, beats_per_measure: int, leftover_ticks: int, beat_ticks: int) -> str:
        if leftover_ticks == 0:
            return f"{beat_number}-of-{beats_per_measure}"
        fraction = Fraction(leftover_ticks, beat_ticks).limit_denominator(8)
        if fraction == 0:
            return f"{beat_number}-of-{beats_per_measure}"
        return f"{beat_number}+{fraction}-of-{beats_per_measure}"


def _normalise_track_name(name: str) -> str:
    return name.strip().lower()


def _build_part_track_map(midi_file: mido.MidiFile) -> Dict[str, mido.MidiTrack]:
    """Locate the MIDI tracks that correspond to the nine Light Chorus parts."""

    name_to_track: Dict[str, mido.MidiTrack] = {}
    for track in midi_file.tracks:
        if track.name:
            name_to_track[_normalise_track_name(track.name)] = track

    part_tracks: Dict[str, mido.MidiTrack] = {}
    for part_name, _ in PART_DEFINITIONS:
        key = _normalise_track_name(part_name)
        if key in name_to_track:
            part_tracks[part_name] = name_to_track[key]
    # Fallback to canonical indices if matching by name failed
    if len(part_tracks) < len(PART_DEFINITIONS):
        indices = list(range(4, 13))
        for (part_name, _), index in zip(PART_DEFINITIONS, indices):
            part_tracks.setdefault(part_name, midi_file.tracks[index])
    return part_tracks


def _collect_note_on_events(track: mido.MidiTrack) -> Dict[int, List[int]]:
    """Return a mapping of absolute tick -> list of MIDI note numbers."""

    tick_map: Dict[int, List[int]] = {}
    abs_tick = 0
    for message in track:
        abs_tick += message.time
        if message.type == "note_on" and message.velocity > 0:
            tick_map.setdefault(abs_tick, []).append(message.note)
    return tick_map


def _note_to_name(note: int, octave_offset: int) -> str:
    pitch_classes = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
    base_name = pitch_classes[note % 12]
    note_name = PREFERRED_ACCIDENTAL_MAP.get(base_name, base_name)
    octave = (note // 12) + octave_offset
    return f"{note_name}{octave}"


def _note_to_short_index(note: int, short_base_note: int) -> int:
    index = note - short_base_note
    if index < 0:
        raise ValueError(f"Note {note} is below available short sample range")
    return index


def _build_note_entry(note: int, octave_offset: int, short_base_note: int) -> NoteEntry:
    short_index = _note_to_short_index(note, short_base_note)
    return NoteEntry(midi=note, note_name=_note_to_name(note, octave_offset), short_sample=str(short_index))


@dataclass
class ProcessingOptions:
    """Configuration flags for MIDI → spreadsheet conversion."""

    octave_offset: int = -1  # scientific pitch: MIDI 60 -> C4
    short_sample_base_note: int = SHORT_SAMPLE_BASE_NOTE


def extract_light_chorus_events(midi_path: str, options: Optional[ProcessingOptions] = None) -> Tuple[List[EventColumn], List[str]]:
    """Parse the MIDI file and return structured Light Chorus event data."""

    opts = options or ProcessingOptions()
    midi_file = mido.MidiFile(midi_path)
    time_mapper = TimeSignatureMap(midi_file)
    part_tracks = _build_part_track_map(midi_file)
    part_tick_maps: Dict[str, Dict[int, List[int]]] = {
        part_name: _collect_note_on_events(track)
        for part_name, track in part_tracks.items()
    }

    unique_ticks = sorted({tick for mapping in part_tick_maps.values() for tick in mapping})
    events: List[EventColumn] = []
    for column_index, tick in enumerate(unique_ticks, start=1):
        measure, beat_label = time_mapper.measure_position(tick)
        part_entries: Dict[str, List[NoteEntry]] = {}
        for part_name, tick_map in part_tick_maps.items():
            notes = tick_map.get(tick)
            if not notes:
                continue
            part_entries[part_name] = [
                _build_note_entry(note, opts.octave_offset, opts.short_sample_base_note)
                for note in notes
            ]
        events.append(EventColumn(number=column_index, tick=tick, measure=measure, position_label=beat_label, part_notes=part_entries))
    part_order = [part_name for part_name, _ in PART_DEFINITIONS]
    return events, part_order


def build_workbook(events: Sequence[EventColumn], part_order: Sequence[str]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Light Chorus"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="000000")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header rows labels
    sheet.cell(row=1, column=1, value="Event #").font = header_font
    sheet.cell(row=1, column=1).fill = header_fill
    sheet.cell(row=2, column=1, value="Measure #").font = header_font
    sheet.cell(row=2, column=1).fill = header_fill
    sheet.cell(row=3, column=1, value="Position (beat)").font = header_font
    sheet.cell(row=3, column=1).fill = header_fill

    for col_index, event in enumerate(events, start=2):
        sheet.cell(row=1, column=col_index, value=event.number)
        sheet.cell(row=2, column=col_index, value=event.measure)
        sheet.cell(row=3, column=col_index, value=event.position_label)
        sheet.cell(row=1, column=col_index).font = header_font
        sheet.cell(row=1, column=col_index).fill = header_fill
        sheet.cell(row=2, column=col_index).font = header_font
        sheet.cell(row=2, column=col_index).fill = header_fill
        sheet.cell(row=3, column=col_index).font = header_font
        sheet.cell(row=3, column=col_index).fill = header_fill
        sheet.cell(row=1, column=col_index).alignment = alignment_center
        sheet.cell(row=2, column=col_index).alignment = alignment_center
        sheet.cell(row=3, column=col_index).alignment = alignment_center

    color_lookup = {name: color for name, color in PART_DEFINITIONS}

    # Part rows with styling (two rows per part: primer tone paths then pitch names)
    for part_index, part_name in enumerate(part_order):
        base_row = 4 + part_index * 2
        primer_row = base_row
        pitch_row = base_row + 1
        color_hex = color_lookup.get(part_name, "#3c3c3c").replace("#", "")

        for row, label in ((primer_row, part_name), (pitch_row, "")):
            label_cell = sheet.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True, color="FFFFFF")
            label_cell.fill = PatternFill(fill_type="solid", fgColor=color_hex)
            label_cell.alignment = alignment_center

        for col_index, event in enumerate(events, start=2):
            entries = event.part_notes.get(part_name)
            primer_cell = sheet.cell(row=primer_row, column=col_index)
            pitch_cell = sheet.cell(row=pitch_row, column=col_index)
            if entries:
                primer_cell.value = ", ".join(entry.sample_path for entry in entries)
                pitch_cell.value = ", ".join(entry.note_name for entry in entries)
            primer_cell.alignment = alignment_center
            pitch_cell.alignment = alignment_center

    # Adjust column widths for readability
    for col_index in range(1, len(events) + 2):
        column_letter = get_column_letter(col_index)
        if col_index == 1:
            sheet.column_dimensions[column_letter].width = 18
        else:
            sheet.column_dimensions[column_letter].width = 20

    sheet.freeze_panes = "B4"
    return workbook


def process_midi_to_workbook(midi_path: str, output_path: str, options: Optional[ProcessingOptions] = None) -> Workbook:
    events, part_order = extract_light_chorus_events(midi_path, options)
    workbook = build_workbook(events, part_order)
    workbook.save(output_path)
    return workbook
